from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from packages.core.domain.models import hash_text
from packages.core.services.company_normalize import canonical_company_display
from packages.core.services.job_intake import ImportResult, upsert_job_from_import

_ABOUT_RE = re.compile(r"^about\s+(.+)$", re.IGNORECASE)
_LOCATION_LIKE = re.compile(r",\s*[A-Z]{2}\b")


def _looks_like_location(val: str | None) -> bool:
    if not val:
        return False
    return bool(_LOCATION_LIKE.search(val)) and len(val) < 80


def _infer_company_from_description(parts: list[str]) -> str | None:
    for part in parts:
        m = _ABOUT_RE.match(part.strip())
        if m:
            return canonical_company_display(m.group(1).strip())
    for part in parts:
        if "globalfoundries" in part.lower():
            return "GlobalFoundries"
    return None


def _read_symplicity_jobs(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header = [str(c).strip() if c else "" for c in rows[0]]
    idx = {name: i for i, name in enumerate(header)}

    jobs: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    desc_parts: list[str] = []

    def flush() -> None:
        nonlocal current, desc_parts
        if not current:
            return
        description = "\n".join(desc_parts).strip()
        if description:
            current["description_text"] = description
        if not current.get("company"):
            inferred = _infer_company_from_description(desc_parts)
            if inferred:
                current["company"] = inferred
        if current.get("title") and current.get("company") and current.get("description_text"):
            jobs.append(current)
        current = None
        desc_parts = []

    for row in rows[1:]:
        row_list = list(row)
        if not any(row_list):
            continue
        source_val = row_list[idx.get("Source", 0)] if idx.get("Source") is not None else None
        if source_val and str(source_val).strip().lower() == "symplicity":
            flush()
            title = row_list[idx["Title"]] if "Title" in idx else None
            company = row_list[idx["Company"]] if "Company" in idx else None
            location = row_list[idx["Location"]] if "Location" in idx else None
            posted_at = row_list[idx["Posted At"]] if "Posted At" in idx else None
            # Sample sheet stores Dallas, TX under Posted At when Location empty
            if not location and _looks_like_location(str(posted_at or "")):
                location = posted_at
                posted_at = None
            desc_cell = row_list[idx["Description"]] if "Description" in idx else None
            current = {
                "title": str(title).strip() if title else "",
                "company": canonical_company_display(str(company)) if company else "",
                "location": str(location).strip() if location else None,
                "posted_at": posted_at,
                "canonical_url": row_list[idx.get("URL", 1)] if "URL" in idx else None,
                "external_id": row_list[idx.get("External ID", 2)] if "External ID" in idx else None,
                "notes": row_list[idx.get("Notes", 8)] if "Notes" in idx else None,
            }
            if desc_cell:
                desc_parts.append(str(desc_cell))
            ext = current.get("external_id")
            current["source_import_key"] = hash_text(
                f"symplicity|{ext or current.get('canonical_url') or current.get('title')}"
            )
            continue
        if current is not None:
            desc_idx = idx.get("Description", 7)
            if desc_idx < len(row_list) and row_list[desc_idx]:
                desc_parts.append(str(row_list[desc_idx]))

    flush()
    return jobs


def import_symplicity_xlsx(session: Session, path: str) -> ImportResult:
    res = ImportResult()
    p = Path(path)
    if not p.is_file():
        res.errors.append(f"file not found: {path}")
        return res
    for job in _read_symplicity_jobs(p):
        if not job.get("company"):
            res.errors.append(f"missing company for title={job.get('title')}")
            continue
        payload = {
            **job,
            "description_text": job.get("description_text") or "",
            "intake_metadata": {"symplicity_notes": job.get("notes")},
            "raw_payload": job,
        }
        _, created = upsert_job_from_import(
            session,
            payload,
            source_name="symplicity",
            source_type="xlsx_import",
            legacy_source="json_import",
        )
        if created:
            res.created += 1
        else:
            res.duplicates += 1
    return res
