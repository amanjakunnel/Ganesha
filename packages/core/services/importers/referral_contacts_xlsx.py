from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from packages.core.domain.models import ReferralContact, hash_text
from packages.core.services.company_normalize import canonical_company_display, normalize_company_key
from packages.core.services.job_intake import ImportResult, get_or_create_company


def import_referral_contacts_xlsx(session: Session, path: str) -> ImportResult:
    res = ImportResult()
    p = Path(path)
    if not p.is_file():
        res.errors.append(f"file not found: {path}")
        return res

    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return res
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return res
    header = [str(c).strip() if c else "" for c in rows[0]]
    idx = {name: i for i, name in enumerate(header)}

    for i, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue
        def cell(name: str, default: int) -> Any:
            i = idx.get(name, default)
            if i is None or i >= len(row):
                return None
            return row[i]

        company_raw = cell("Company", 0)
        name = cell("Name", 1)
        if not company_raw or not name:
            if not any(row):
                continue
            continue
        company_name = canonical_company_display(str(company_raw))
        contact_name = str(name).strip()
        position = cell("Position", 2)
        team = cell("Team", 3)
        locations = cell("Location(s)", 4)
        alt = cell("Alternate Location", 5)
        import_key = hash_text(
            f"referral|{normalize_company_key(company_name)}|{contact_name.lower()}"
        )
        existing = (
            session.query(ReferralContact)
            .filter(ReferralContact.source_import_key == import_key)
            .one_or_none()
        )
        company = get_or_create_company(session, company_name)
        if existing:
            existing.position = str(position) if position else existing.position
            existing.team = str(team) if team else existing.team
            existing.locations = str(locations) if locations else existing.locations
            existing.alternate_location = str(alt) if alt else existing.alternate_location
            existing.company_id = company.id if company else existing.company_id
            session.add(existing)
            res.updated += 1
            continue
        rc = ReferralContact(
            company_id=company.id if company else None,
            company_name_raw=company_name,
            contact_name=contact_name,
            position=str(position) if position else None,
            team=str(team) if team else None,
            locations=str(locations) if locations else None,
            alternate_location=str(alt) if alt else None,
            source_import_key=import_key,
        )
        session.add(rc)
        res.created += 1
    session.flush()
    return res
